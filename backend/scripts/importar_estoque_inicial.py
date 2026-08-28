"""Importação de carga inicial de estoque a partir de planilha (.xlsx/.csv).

Ferramenta de uso único, pensada para popular o catálogo de medicamentos e
o estoque físico (lotes) na instalação nova de um hospital, a partir de uma
planilha de conferência já existente — evita ter que cadastrar item por
item pela tela. Não é uma tela/endpoint do sistema (fora do MVP, mesmo
espírito de scripts/seed_usuarios.py): roda uma vez, direto no banco.

Como é carga inicial, este script grava lotes em QUALQUER unidade (CAF,
UTI, Centro Cirúrgico, Emergência) — diferente da regra normal do sistema
("Entrada só ocorre na CAF", docs/00_PROJETO.md seção 3), que só vale para
o fluxo do dia a dia pela tela/API. Aqui estamos registrando o que já
existe fisicamente em cada unidade no momento zero, não uma entrada nova.

Colunas esperadas na planilha (cabeçalho na primeira linha, nomes exatos
abaixo — ver docs/IMPORTACAO_ESTOQUE_INICIAL.md para a explicação de cada
uma e os valores aceitos):

    medicamento, apresentacao, concentracao, fabricante, acondicionamento,
    estoque_minimo, unidade, numero_lote, data_validade, quantidade,
    valor_unitario, origem, numero_nota_fiscal, numero_afm,
    e_antimicrobiano, e_controlado

Uso:

    # 1. Sempre rodar em modo teste primeiro (não grava nada no banco):
    python scripts/importar_estoque_inicial.py --arquivo estoque.xlsx --usuario-login ananda.carvalho

    # 2. Só depois de revisar os erros/avisos, gravar de verdade:
    python scripts/importar_estoque_inicial.py --arquivo estoque.xlsx --usuario-login ananda.carvalho --confirmar
"""

import argparse
import csv
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.database.database import SessionLocal  # noqa: E402
from app.models.enums import ApresentacaoEnum, AcondicionamentoEnum, OrigemEnum, TipoMovimentacaoEnum  # noqa: E402
from app.models.lote import Lote  # noqa: E402
from app.models.medicamento import Medicamento  # noqa: E402
from app.models.movimentacao import Movimentacao  # noqa: E402
from app.models.unidade import Unidade  # noqa: E402
from app.models.usuario import Usuario  # noqa: E402

COLUNAS_OBRIGATORIAS = [
    "medicamento",
    "apresentacao",
    "concentracao",
    "acondicionamento",
    "unidade",
    "numero_lote",
    "data_validade",
    "quantidade",
    "origem",
]

VALORES_SIM = {"sim", "s", "true", "1", "verdadeiro", "x"}


@dataclass
class LinhaImportada:
    numero_linha: int
    dados: dict
    erros: list[str] = field(default_factory=list)


def ler_planilha(caminho: Path) -> list[dict]:
    if caminho.suffix.lower() == ".csv":
        with open(caminho, newline="", encoding="utf-8-sig") as f:
            return list(csv.DictReader(f))

    if caminho.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        wb = load_workbook(caminho, read_only=True, data_only=True)
        ws = wb.active

        linhas_brutas = list(ws.iter_rows(values_only=True))
        if not linhas_brutas:
            return []

        cabecalho = [str(c).strip() if c is not None else "" for c in linhas_brutas[0]]
        registros = []
        for linha in linhas_brutas[1:]:
            if all(v is None for v in linha):
                continue
            registros.append(dict(zip(cabecalho, linha)))
        return registros

    raise SystemExit(f"Formato de arquivo não suportado: {caminho.suffix} (use .xlsx ou .csv)")


def texto(valor) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def parse_data(valor) -> date | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    valor = str(valor).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(valor, fmt).date()
        except ValueError:
            continue
    return None


def parse_decimal(valor) -> Decimal | None:
    if valor is None or valor == "":
        return Decimal("0")
    if isinstance(valor, (int, float)):
        return Decimal(str(valor))
    valor = str(valor).strip().replace("R$", "").strip()
    valor = valor.replace(".", "").replace(",", ".") if "," in valor else valor
    try:
        return Decimal(valor)
    except InvalidOperation:
        return None


def parse_inteiro(valor) -> int | None:
    if valor is None or valor == "":
        return None
    try:
        return int(float(str(valor).strip().replace(",", ".")))
    except ValueError:
        return None


def parse_bool(valor) -> bool:
    return texto(valor).lower() in VALORES_SIM


def validar_linha(linha: LinhaImportada, unidades_por_nome: dict[str, Unidade]) -> None:
    d = linha.dados

    for coluna in COLUNAS_OBRIGATORIAS:
        if not texto(d.get(coluna)):
            linha.erros.append(f"coluna '{coluna}' vazia/ausente")

    if linha.erros:
        return  # sem as colunas básicas, não adianta validar o resto

    apresentacao_raw = texto(d["apresentacao"]).lower()
    if apresentacao_raw not in ApresentacaoEnum.__members__:
        linha.erros.append(
            f"apresentacao '{d['apresentacao']}' inválida — valores aceitos: "
            + ", ".join(ApresentacaoEnum.__members__)
        )

    acond_raw = texto(d["acondicionamento"]).lower()
    if acond_raw not in AcondicionamentoEnum.__members__:
        linha.erros.append(
            f"acondicionamento '{d['acondicionamento']}' inválido — use 'ambiente' ou 'geladeira'"
        )

    unidade_raw = texto(d["unidade"])
    if unidade_raw not in unidades_por_nome:
        nomes_unidades = sorted(u.nome for u in unidades_por_nome.values() if u.tipo == "unidade")
        linha.erros.append(
            f"unidade '{unidade_raw}' não encontrada — unidades reais: "
            + ", ".join(nomes_unidades)
            + ". Pra carrinho, use o nome exato cadastrado (ver tela Reposição de Carrinhos)."
        )

    origem_raw = texto(d["origem"]).lower()
    if origem_raw not in OrigemEnum.__members__:
        linha.erros.append(f"origem '{d['origem']}' inválida — use 'compra' ou 'doacao'")

    data_validade = parse_data(d.get("data_validade"))
    if data_validade is None:
        linha.erros.append(
            f"data_validade '{d.get('data_validade')}' não reconhecida — use DD/MM/AAAA"
        )

    quantidade = parse_inteiro(d.get("quantidade"))
    if quantidade is None or quantidade <= 0:
        linha.erros.append("quantidade precisa ser um número inteiro maior que zero")

    valor_unitario = parse_decimal(d.get("valor_unitario"))
    if valor_unitario is None:
        linha.erros.append(f"valor_unitario '{d.get('valor_unitario')}' não reconhecido")

    if origem_raw == "compra":
        if not texto(d.get("numero_nota_fiscal")):
            linha.erros.append("numero_nota_fiscal é obrigatório quando origem=compra")
        if valor_unitario is not None and valor_unitario <= 0:
            linha.erros.append("valor_unitario deve ser maior que zero quando origem=compra")

    estoque_minimo = d.get("estoque_minimo")
    if texto(estoque_minimo) and parse_inteiro(estoque_minimo) is None:
        linha.erros.append(f"estoque_minimo '{estoque_minimo}' não é um número inteiro válido")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--arquivo", required=True, help="Caminho da planilha (.xlsx ou .csv)")
    parser.add_argument(
        "--usuario-login",
        required=True,
        help="Login de um usuário já cadastrado (farmacêutico/coordenador) — fica registrado como autor da entrada na trilha de auditoria",
    )
    parser.add_argument(
        "--confirmar",
        action="store_true",
        help="Sem essa flag, roda em modo teste (só valida, não grava nada no banco)",
    )
    args = parser.parse_args()

    caminho = Path(args.arquivo)
    if not caminho.exists():
        raise SystemExit(f"Arquivo não encontrado: {caminho}")

    registros = ler_planilha(caminho)
    if not registros:
        raise SystemExit("Planilha vazia ou sem linhas de dados.")

    db = SessionLocal()
    try:
        usuario = db.query(Usuario).filter(Usuario.login == args.usuario_login).first()
        if usuario is None:
            raise SystemExit(f"Usuário '{args.usuario_login}' não encontrado.")

        # Inclui unidades reais E carrinhos de emergência — um lote de carga
        # inicial pode estar fisicamente num carrinho, não só numa unidade
        # (os 18 carrinhos já existem desde a migration 0003, cada um com
        # tipo="carrinho" e unidade_pai_id apontando pra unidade real onde
        # fica). Na planilha, usar o nome exato do carrinho (ex.: "Carro de
        # Emergência Unidade Canguru"), igual aparece na tela Reposição de
        # Carrinhos — não precisa citar a unidade-mãe à parte.
        unidades = db.query(Unidade).all()
        unidades_por_nome = {u.nome: u for u in unidades}

        linhas = [
            LinhaImportada(numero_linha=i + 2, dados=r) for i, r in enumerate(registros)
        ]
        for linha in linhas:
            validar_linha(linha, unidades_por_nome)

        linhas_com_erro = [l for l in linhas if l.erros]
        linhas_ok = [l for l in linhas if not l.erros]

        print(f"Total de linhas: {len(linhas)} | válidas: {len(linhas_ok)} | com erro: {len(linhas_com_erro)}\n")

        if linhas_com_erro:
            print("=== Linhas com erro (não serão importadas) ===")
            for linha in linhas_com_erro:
                nome = texto(linha.dados.get("medicamento")) or "(sem nome)"
                print(f"  Linha {linha.numero_linha} ({nome}):")
                for erro in linha.erros:
                    print(f"    - {erro}")
            print()

        if not args.confirmar:
            print("Modo TESTE (nada foi gravado). Corrija os erros acima e rode de novo.")
            print("Quando a planilha estiver limpa, rode com --confirmar para gravar de verdade.")
            return

        if not linhas_ok:
            print("Nenhuma linha válida para importar.")
            return

        medicamentos_cache: dict[tuple[str, str, str], Medicamento] = {}
        criados_medicamento = 0
        criados_lote = 0

        for linha in linhas_ok:
            d = linha.dados
            nome = texto(d["medicamento"])
            apresentacao = texto(d["apresentacao"]).lower()
            concentracao = texto(d["concentracao"])
            chave = (nome.lower(), apresentacao, concentracao.lower())

            medicamento = medicamentos_cache.get(chave)
            if medicamento is None:
                medicamento = (
                    db.query(Medicamento)
                    .filter(
                        Medicamento.nome.ilike(nome),
                        Medicamento.apresentacao == apresentacao,
                        Medicamento.concentracao.ilike(concentracao),
                    )
                    .first()
                )

            if medicamento is None:
                medicamento = Medicamento(
                    nome=nome,
                    apresentacao=apresentacao,
                    concentracao=concentracao,
                    fabricante=texto(d.get("fabricante")) or None,
                    acondicionamento=texto(d["acondicionamento"]).lower(),
                    estoque_minimo=parse_inteiro(d.get("estoque_minimo")) or 0,
                    e_antimicrobiano=parse_bool(d.get("e_antimicrobiano")),
                    e_controlado=parse_bool(d.get("e_controlado")),
                )
                db.add(medicamento)
                db.flush()  # garante medicamento.id sem commitar ainda
                criados_medicamento += 1

            medicamentos_cache[chave] = medicamento

            unidade = unidades_por_nome[texto(d["unidade"])]
            origem_raw = texto(d["origem"]).lower()
            valor_unitario = parse_decimal(d.get("valor_unitario")) or Decimal("0")
            if origem_raw in ("doacao", "emprestimo"):
                valor_unitario = Decimal("0")

            lote = Lote(
                medicamento_id=medicamento.id,
                unidade_id=unidade.id,
                numero_lote=texto(d["numero_lote"]),
                data_validade=parse_data(d["data_validade"]),
                quantidade_atual=parse_inteiro(d["quantidade"]),
                valor_unitario=valor_unitario,
                origem=origem_raw,
                numero_nota_fiscal=texto(d.get("numero_nota_fiscal")) or None,
                numero_afm=texto(d.get("numero_afm")) or None,
                usuario_entrada_id=usuario.id,
            )
            db.add(lote)
            db.flush()

            movimentacao = Movimentacao(
                tipo=TipoMovimentacaoEnum.entrada,
                lote_id=lote.id,
                quantidade=lote.quantidade_atual,
                unidade_destino_id=unidade.id,
                usuario_id=usuario.id,
            )
            db.add(movimentacao)
            criados_lote += 1

        db.commit()
        print(f"Importação concluída: {criados_medicamento} medicamento(s) novo(s), {criados_lote} lote(s) criado(s).")

    finally:
        db.close()


if __name__ == "__main__":
    main()
