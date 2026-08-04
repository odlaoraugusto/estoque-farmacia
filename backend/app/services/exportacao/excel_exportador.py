"""Gera o `.xlsx` de uma `TabelaRelatorio` — uma planilha por relatório,
com o cabeçalho institucional (docs/00_PROJETO.md seção 14) mesclado nas
primeiras linhas, antes da tabela de dados."""

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.exportacao.formatacao import formatar_data_hora
from app.services.exportacao.tabela import TabelaRelatorio

_COR_CABECALHO = "2F5233"
_LARGURA_COLUNA_PADRAO = 22
_NOME_PLANILHA_INVALIDO = re.compile(r"[\[\]:*?/\\]")
_GATILHOS_FORMULA = ("=", "+", "-", "@")


def _valor_seguro(valor):
    """Mitigação de 'CSV/Excel injection': texto livre do usuário (nome de
    medicamento, setor consumidor, motivo de descarte) que comece com
    =, +, - ou @ seria interpretado como fórmula ao abrir no Excel.
    Prefixar com `'` é o próprio escape nativo do Excel para "isto é
    texto literal" — não aparece na célula renderizada."""
    if isinstance(valor, str) and valor.startswith(_GATILHOS_FORMULA):
        return "'" + valor
    return valor


class ExcelExportador:
    def gerar(self, tabela: TabelaRelatorio) -> bytes:
        workbook = Workbook()
        planilha = workbook.active
        planilha.title = self._nome_planilha(tabela.metadados.titulo_relatorio)

        num_colunas = max(len(tabela.colunas), 1)
        linha_atual = self._escrever_letterhead(planilha, tabela, num_colunas)
        linha_atual = self._escrever_tabela(planilha, tabela, linha_atual)
        self._escrever_rodape(planilha, tabela, linha_atual, num_colunas)
        self._ajustar_largura_colunas(planilha, num_colunas)

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _nome_planilha(titulo: str) -> str:
        nome = _NOME_PLANILHA_INVALIDO.sub("", titulo).strip()
        return (nome or "Relatório")[:31]

    def _escrever_letterhead(
        self, planilha: Worksheet, tabela: TabelaRelatorio, num_colunas: int
    ) -> int:
        linha = 1
        metadados = tabela.metadados

        linha = self._linha_mesclada(planilha, linha, num_colunas, metadados.organizacao, 13)
        linha = self._linha_mesclada(planilha, linha, num_colunas, metadados.hospital, 12)
        linha = self._linha_mesclada(planilha, linha, num_colunas, metadados.titulo_relatorio, 12)
        linha = self._linha_mesclada(
            planilha,
            linha,
            num_colunas,
            (
                f"Gerado em {formatar_data_hora(metadados.gerado_em)} · "
                f"{metadados.gerado_por} · Unidade: {metadados.unidade}"
            ),
            9,
            negrito=False,
            italico=True,
        )

        for info in tabela.informacoes_extra:
            linha = self._linha_mesclada(planilha, linha, num_colunas, info, 10, negrito=False)

        return linha + 1  # linha em branco antes da tabela

    @staticmethod
    def _linha_mesclada(
        planilha: Worksheet,
        linha: int,
        num_colunas: int,
        texto: str,
        tamanho: int,
        negrito: bool = True,
        italico: bool = False,
    ) -> int:
        celula = planilha.cell(row=linha, column=1, value=texto)
        planilha.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=num_colunas)
        celula.font = Font(bold=negrito, italic=italico, size=tamanho)
        celula.alignment = Alignment(horizontal="left", vertical="center")
        return linha + 1

    @staticmethod
    def _escrever_tabela(planilha: Worksheet, tabela: TabelaRelatorio, linha_inicial: int) -> int:
        for indice, coluna in enumerate(tabela.colunas, start=1):
            celula = planilha.cell(row=linha_inicial, column=indice, value=coluna)
            celula.font = Font(bold=True, color="FFFFFF")
            celula.fill = PatternFill("solid", fgColor=_COR_CABECALHO)
            celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        linha = linha_inicial + 1
        for dados_linha in tabela.linhas:
            for indice, valor in enumerate(dados_linha, start=1):
                planilha.cell(row=linha, column=indice, value=_valor_seguro(valor))
            linha += 1

        return linha

    @staticmethod
    def _escrever_rodape(
        planilha: Worksheet, tabela: TabelaRelatorio, linha: int, num_colunas: int
    ) -> None:
        if not tabela.rodape:
            return

        linha += 1
        for texto in tabela.rodape:
            celula = planilha.cell(row=linha, column=1, value=texto)
            planilha.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=num_colunas)
            celula.font = Font(bold=True, size=11)
            linha += 1

    @staticmethod
    def _ajustar_largura_colunas(planilha: Worksheet, num_colunas: int) -> None:
        for indice in range(1, num_colunas + 1):
            planilha.column_dimensions[get_column_letter(indice)].width = _LARGURA_COLUNA_PADRAO
